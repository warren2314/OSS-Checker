import os
import time
import requests
import re
import logging
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment


def extract_name_and_version_v2(filename):
    base_name = filename.split('.conda')[0].split('.tar.bz2')[0]
    parts = base_name.split('-')
    name = '-'.join(parts[:-2])
    version = parts[-2]
    return f"{name}@{version}"


def process_conda_directory(directory_path):
    if not os.path.exists(directory_path):
        print(f"Directory '{directory_path}' does not exist!")
        return []

    package_details = ["Ecosystem: conda"]

    for root, dirs, files in os.walk(directory_path):
        for file in files:
            if file.endswith('.conda') or file.endswith('.tar.bz2'):
                result = extract_name_and_version_v2(file)
                print(result)
                package_details.append(result)

    return package_details


def get_package_details(filename):
    match = re.match(r'([a-zA-Z0-9_]+)-([\d.]+)', filename)
    if match:
        package_name = match.group(1)
        version = match.group(2)
        return f"{package_name}@{version}"
    else:
        return None


def process_python_directory(directory_path, output_filename):
    package_details = ["Ecosystem: pypi"]

    with open(output_filename, 'w') as output_file:
        for root, dirs, files in os.walk(directory_path):
            for file in files:
                if file.endswith('.whl'):
                    package_detail = get_package_details(file)
                    if package_detail:
                        output_file.write(package_detail + '\n')
                        print(package_detail)
                        package_details.append(package_detail)

    return package_details


# Function to process NuGet directory
def process_nuget_directory(directory_path):
    # Initialize a list with the ecosystem information
    package_details = ["Ecosystem: nuget"]
    
    for root, dirs, files in os.walk(directory_path):
        # Check if the current directory contains .nupkg files
        if any(file.endswith('.nupkg') for file in files):
            # Extract the package name and version from the directory structure
            parts = root.split(os.sep)
            if len(parts) > 2:
                version = parts[-1]  # The version is the last directory
                package_name = parts[-2]  # The package name is one directory up
                package_details.append(f"{package_name}@{version}")
                
    return package_details


# Process Maven packages
def process_maven_directory(directory_path):
    # Initialize a list with the ecosystem information
    artifact_version_data = ["Ecosystem: maven"]
    for root, dirs, files in os.walk(directory_path):
        # Check if the current directory contains .pom or .jar files
        if any(file.endswith(('.pom', '.jar')) for file in files):
            parts = root.split(os.sep)
            # Ensure there are enough parts to get artifact and version
            if len(parts) > 3:
                version = parts[-2]  # The version is two directories up
                artifact = parts[-3]  # The artifact name is three directories up
                artifact_version_data.append(f"{artifact}@{version}")
                #print(f"Debug: Appending artifact: {artifact}@{version}")  # Debug statement
    # Remove duplicates from the artifact and version data, keeping the ecosystem information intact
    artifact_version_data = [artifact_version_data[0]] + list(set(artifact_version_data[1:]))
    return artifact_version_data


def extract_artifacts_to_file(directory_path, structure_type):
    artifact_version_data = []

    if structure_type == "conda":
        artifact_version_data.extend(process_conda_directory(directory_path))
    elif structure_type == "pypi":
        output_filename = os.path.join(directory_path, "python_output.txt")
        artifact_version_data.extend(process_python_directory(directory_path, output_filename))
    elif structure_type == "rpm":  # Add this new case for RPM
        artifact_version_data.extend(process_rpm_directory(directory_path))
    elif structure_type == "maven":
        artifact_version_data.extend(process_maven_directory(directory_path))
    elif structure_type == "nuget":
    	artifact_version_data.extend(process_nuget_directory(directory_path))
    else:
        print("Unsupported structure type.")
        return

    return artifact_version_data, structure_type


def get_vulnerabilities(chunk, credentials, ecosystem):
    url = "https://ossindex.sonatype.org/api/v3/component-report"
    results = []

    for package in chunk:
        if '@' in package:  # Check if the package string contains a version
            package_name, package_version = package.split('@')
            coordinate = f"pkg:{ecosystem}/{package_name}@{package_version}"
        else:
            package_name = package
            coordinate = f"pkg:{ecosystem}/{package_name}"

        payload = {"coordinates": [coordinate]}

        headers = {
            'Content-Type': 'application/json',
            'Authorization': f'Basic {credentials}'
        }

        try:
            response = requests.post(url, json=payload, headers=headers)
            if response.status_code != 200:
                print(f"Error: Received status code {response.status_code} from the API for package {package}")
                continue  # Skip this package and continue with the next one

            result = response.json()[0]
            print(result)
            results.append(result)

        except requests.exceptions.RequestException as e:
            print(f"Error occurred while processing package {package}: {e}")
            continue  # Skip this package and continue with the next one

    return results


def main():
    load_dotenv()
    credentials = os.getenv('API_KEY')
    print(f"Credentials: {credentials}")

    directory_path = input("Enter the directory path: ")
    structure_types = input("Enter the structure types (conda, pypi, rpm, maven) separated by commas: ").lower().split(
        ',')

    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    for structure_type in structure_types:
        structure_type = structure_type.strip()
        print(f"Processing {structure_type} structure...")

        artifact_version_data, ecosystem = extract_artifacts_to_file(directory_path, structure_type)

        with open(f'oss_index_{ecosystem}.txt', 'w') as f:
            for line in artifact_version_data:
                f.write(line + "\n")

        print(f"The text file for {ecosystem} has been created.")

        with open(f'oss_index_{ecosystem}.txt', 'r') as input_file:
            ecosystem_line = input_file.readline().strip()
            print(f"Debug: ecosystem_line = {ecosystem_line}")
            packages = []  # Initialize packages here
            if ": " in ecosystem_line:  # Add this check
                ecosystem = ecosystem_line.split(": ")[1]
            else:
                print(f"Unexpected format: {ecosystem_line}")
                ecosystem = "unknown"
            print(f"Identified Ecosystem: {ecosystem}")
            packages.extend(line.strip() for line in input_file)

        ws = wb.create_sheet(title=f"{ecosystem.capitalize()} Vulnerabilities")
        row = 1

        ws.cell(row=row, column=1, value="Title")
        ws.cell(row=row, column=2, value="Score")
        ws.cell(row=row, column=3, value="CVE")
        ws.cell(row=row, column=4, value="Description")

        row += 1

        calls_per_minute = 120
        seconds_between_calls = 60 / calls_per_minute
        chunk_size = 128

        for i in range(0, len(packages), chunk_size):
            chunk = packages[i:i + chunk_size]
            print(f"Checking package(s): {', '.join(chunk)}")
            try:
                results = get_vulnerabilities(chunk, credentials, ecosystem)
            except requests.exceptions.RequestException as e:
                print(f"Error occurred while processing chunk: {e}")
                continue

            for result in results:
                if result['vulnerabilities']:
                    print(f"{result['coordinates']}: {len(result['vulnerabilities'])} known vulnerabilities")
                    for vulnerability in result['vulnerabilities']:
                        cve = vulnerability.get('cve', 'N/A')
                        title = vulnerability.get('title', 'N/A')
                        cvss_score = vulnerability.get('cvssScore', 'N/A')
                        description = vulnerability.get('description', 'N/A')

                        ws.cell(row=row, column=1, value=f"{title}")
                        ws.cell(row=row, column=2, value=f"{cvss_score}")
                        ws.cell(row=row, column=3, value=f"{cve}")
                        ws.cell(row=row, column=4, value=f"{description}")

                        for col in range(1, 5):
                            ws.cell(row=row, column=col).alignment = Alignment(wrapText=True, vertical='top',
                                                                               horizontal='left')

                        row += 1

            time.sleep(seconds_between_calls)

    wo_number = input("Please enter a 6-digit WO number: ")
    while len(wo_number) != 6 or not wo_number.isdigit():
        print("Invalid input. Please ensure you enter a 6-digit number and exclude the WO.")
        wo_number = input("Please enter a 6-digit WO number: ")

    filename = f"WO{wo_number}_vulnerabilities.xlsx"
    wb.save(filename)
    print(f"Vulnerabilities saved to {filename}")


if __name__ == '__main__':
    main()
